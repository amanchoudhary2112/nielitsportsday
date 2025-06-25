from flask import Flask, session, render_template, request, redirect, url_for, flash, send_file
import os
import openpyxl
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'aman_key'

EXCEL_FILE = 'participants.xlsx'

# Ensure Excel file is initialized
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Participants"
    sheet.append(["Name", "Email", "Sport", "Roll No.", "Mobile Number"])
    wb.save(EXCEL_FILE)

# -------------------- ROUTES -------------------- #

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == 'admin' and password == 'admin123':
            session['admin'] = True
            return redirect(url_for('participants'))
        else:
            flash('Invalid credentials', 'error')
    return render_template('login.html')


@app.route('/')
def index():
    sports = {
        'Football': '/static/images/football.jpg',
        'Basketball': '/static/images/basketball.jpg',
        'Cricket': '/static/images/cricket.jpg',
        'Athletics': '/static/images/athletics.jpg'
    }
    return render_template('index.html', sports=sports)


@app.route('/register', defaults={'sport': None}, methods=['GET', 'POST'])
@app.route('/register/<sport>', methods=['GET', 'POST'])
def register(sport):
    sports_list = ['Football', 'Basketball', 'Cricket', 'Athletics']

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        sport_selected = request.form['sport']
        roll_no = request.form.get('roll_no', '')
        mobile = request.form.get('mobile', '')

        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb['Participants']
        sheet.append([name, email, sport_selected, roll_no, mobile])
        wb.save(EXCEL_FILE)

        return redirect(url_for('success', name=name, sport=sport_selected))

    return render_template('register.html', sports=sports_list, selected_sport=sport)


@app.route('/success')
def success():
    name = request.args.get('name')
    sport = request.args.get('sport')
    return render_template('success.html', name=name, sport=sport)


@app.route('/gallery')
def eventgallery():
    return render_template('gallery.html')


@app.route('/participants')
def participants():
    if not session.get('admin'):
        return redirect(url_for('login'))
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['Participants']
    participants = list(sheet.iter_rows(values_only=True))[1:]  # Skip header row

    return render_template('participants.html', participants=participants, current_year=datetime.now().year)


@app.route('/download')
def download_excel():
    if not session.get('admin'):  # Fixed: use 'admin' to match login session
        return redirect(url_for('login'))
    return send_file(EXCEL_FILE, as_attachment=True)


@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('index'))


# -------------------- MAIN -------------------- #

if __name__ == '__main__':
    app.run(debug=True)
